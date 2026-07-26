from __future__ import annotations

import io
from pathlib import Path

from app.core.exceptions import AppError

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".xlsm"}


class DocumentParser:
    """Extract plain text from supported office document formats without writing to disk."""

    def parse(self, filename: str, content: bytes) -> str:
        extension = Path(filename).suffix.lower()
        if extension not in SUPPORTED_EXTENSIONS:
            raise AppError(
                415,
                "UNSUPPORTED_DOCUMENT",
                "Only PDF, Word (.docx), and Excel (.xlsx/.xlsm) files are supported.",
            )
        if extension == ".pdf":
            from pypdf import PdfReader
        elif extension == ".docx":
            from docx import Document
        else:
            from openpyxl import load_workbook

        try:
            if extension == ".pdf":
                text = "\n\n".join(
                    page.extract_text() or "" for page in PdfReader(io.BytesIO(content)).pages
                )
            elif extension == ".docx":
                document = Document(io.BytesIO(content))
                text = "\n".join(paragraph.text for paragraph in document.paragraphs)
            else:
                workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                rows: list[str] = []
                for sheet in workbook.worksheets:
                    rows.append(f"Sheet: {sheet.title}")
                    rows.extend(
                        "\t".join("" if value is None else str(value) for value in row)
                        for row in sheet.iter_rows(values_only=True)
                    )
                text = "\n".join(rows)
        except AppError:
            raise
        except Exception as exc:
            raise AppError(
                422, "DOCUMENT_PARSE_FAILED", f"The document could not be parsed: {exc}"
            ) from exc
        normalized = "\n".join(line.strip() for line in text.splitlines() if line.strip())
        if not normalized:
            raise AppError(422, "EMPTY_DOCUMENT", "The document contains no extractable text.")
        return normalized
